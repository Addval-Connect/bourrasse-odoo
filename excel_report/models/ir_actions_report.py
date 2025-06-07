# -*- coding: utf-8 -*-

import io
import base64
import os
import sys
import subprocess
import re
import logging

from odoo import models, fields, api
import openpyxl


_logger = logging.getLogger(__name__)


class IrActionsReport(models.Model):
    _inherit = 'ir.actions.report'

    def _get_path_source_excel(self):
        if sys.platform == 'win32':
            return r'C:\temp\backups\out.xlsx'
        return '/tmp/backups/out.xlsx'

    def _get_path_convert_excel(self):
        if sys.platform == 'win32':
            return r'C:\temp\backups'
        return '/tmp/backups'

    def _get_path_libreoffice_excel(self):
        # TODO: Provide support for more platforms
        if sys.platform == 'darwin':
            return '/Applications/LibreOffice.app/Contents/MacOS/soffice'
        if sys.platform == 'win32':
            return r"C:\Program Files\LibreOffice\program\soffice.exe"
        return 'libreoffice'

    report_type = fields.Selection(
        selection_add=[('excel', 'EXCEL')], ondelete={'excel': 'cascade'})
    template_excel = fields.Binary(string='Excel template', attachment=True)
    excel_out_report_type = fields.Selection(
        selection=[('excel', 'EXCEL'), ('pdf', 'PDF'),
                   ('ods', 'ODS'), ('odt', 'ODT'), ('html', 'HTML')],
        help="Extension will download users", default='excel')
    excel_path_source = fields.Char(
        string='OS path to source file temporary', default=_get_path_source_excel)
    excel_path_convert_folder = fields.Char(
        string='OS path to converted file temporary', default=_get_path_convert_excel)
    excel_path_libreoffice = fields.Char(
        string='OS path to libreoffice', default=_get_path_libreoffice_excel, help="For linux just libreoffice")

    @api.model
    def render_excel(self, docids, data=None):
        if not data:
            data = {}
        data.setdefault('report_type', 'excel')

        # Get the rendering context - Updated for Odoo 18
        report_model_name = data.get('model')
        if report_model_name and docids:
            # Create context with docs
            docs = self.env[report_model_name].browse(docids)
            data['docs'] = docs
            data['doc_ids'] = docids
            data['doc_model'] = report_model_name
        else:
            data['docs'] = []
            data['doc_ids'] = []
            data['doc_model'] = ''

        # READ DATA
        if not self.template_excel:
            raise ValueError("No Excel template configured for this report")

        content = base64.b64decode(self.template_excel)

        # MERGE DATA
        # open excel sheets
        wb1 = openpyxl.load_workbook(io.BytesIO(content))
        ws1 = wb1.active

        # compare each element
        for doc in data['docs']:
            for row in range(ws1.max_row):
                for column in range(ws1.max_column):
                    val = ws1.cell(row=row + 1, column=column + 1).value
                    if isinstance(val, str):
                        result = re.findall(r"(odoo\(.*?\))$", val)
                        if len(result):
                            try:
                                # Create a safe environment for evaluation
                                eval_context = {
                                    'doc': doc,
                                    'docs': data['docs'],
                                    'user': self.env.user,
                                    'company': self.env.company,
                                    'time': __import__('time'),
                                    'datetime': __import__('datetime'),
                                    'relativedelta': __import__('dateutil.relativedelta').relativedelta,
                                }
                                new_val = eval(result[0][5: -1], eval_context)

                                if isinstance(new_val, float):
                                    new_val = str(new_val).replace('.', ',')
                                elif isinstance(new_val, str):
                                    new_val = str(new_val)
                                else:
                                    # insert image to cell
                                    try:
                                        imgdata = base64.b64decode(new_val)
                                        myio = io.BytesIO(imgdata)
                                        img = openpyxl.drawing.image.Image(myio)
                                        cell = ws1.cell(row=row + 1, column=column + 1)
                                        img.anchor = cell.coordinate
                                        ws1.add_image(img)
                                        new_val = ''
                                    except Exception as error:
                                        _logger.error('Error when trying insert image %s' % error)
                                        new_val = ''
                                ws1.cell(
                                    row=row + 1, column=column + 1).value = re.sub(r"(odoo\(.*?\))$", new_val, val)
                            except Exception as e:
                                _logger.error('Error evaluating expression %s: %s' % (result[0], e))
                                # Keep original value if evaluation fails
                                pass

        # WRITE DATA
        myio = io.BytesIO()
        wb1.save(myio)
        myio.getvalue()
        path_source = self.excel_path_source

        if self.excel_out_report_type != 'excel':
            if not os.path.isdir(self.excel_path_convert_folder):
                os.makedirs(self.excel_path_convert_folder)

            # WRITE DOCX SOURCE
            with open(path_source, 'wb') as f:
                f.write(myio.getbuffer())

            # CONVERT DOCX TO out_report_type
            def convert_to(folder, source, timeout=None):
                args = [self.excel_path_libreoffice, '--headless', '--convert-to', self.excel_out_report_type, '--outdir', folder, source]
                subprocess.run(args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=timeout)

            convert_to(self.excel_path_convert_folder, path_source)

            # READ out_report_type file FROM OS
            myio = io.BytesIO()
            with open(path_source.replace('xlsx', self.excel_out_report_type), 'rb') as fin:
                myio = io.BytesIO(fin.read())

            try:
                os.unlink(path_source)
                os.unlink(path_source.replace(
                    'xlsx', self.excel_out_report_type))
            except (OSError, IOError):
                _logger.error(
                    'Error when trying to remove file %s' % path_source)

            return myio.getvalue(), self.excel_out_report_type
        return myio.getvalue(), 'excel'
